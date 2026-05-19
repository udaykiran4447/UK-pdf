import streamlit as st

st.set_page_config(
    page_title="Compliance Suite",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_bar="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:ital,wght@0,300;0,400;0,500;1,300&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

.stApp {
    background: #0a0a0f;
    min-height: 100vh;
}

/* Hide streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 !important; max-width: 100% !important; }

/* ── Hero ── */
.hero-wrap {
    position: relative;
    overflow: hidden;
    background: linear-gradient(135deg, #0d0d1a 0%, #111128 50%, #0a0f1e 100%);
    padding: 3.5rem 3rem 2.5rem;
    border-bottom: 1px solid rgba(255,255,255,0.07);
}
.hero-wrap::before {
    content: '';
    position: absolute;
    top: -80px; right: -120px;
    width: 500px; height: 500px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(99,102,241,0.18) 0%, transparent 70%);
    pointer-events: none;
}
.hero-wrap::after {
    content: '';
    position: absolute;
    bottom: -60px; left: 40px;
    width: 300px; height: 300px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(16,185,129,0.1) 0%, transparent 70%);
    pointer-events: none;
}
.hero-eyebrow {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: rgba(99,102,241,0.15);
    border: 1px solid rgba(99,102,241,0.35);
    border-radius: 100px;
    padding: 4px 14px;
    font-size: 0.72rem;
    font-weight: 500;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #a5b4fc;
    margin-bottom: 1.2rem;
}
.hero-title {
    font-family: 'Syne', sans-serif;
    font-size: 3rem;
    font-weight: 800;
    line-height: 1.1;
    color: #f0f0ff;
    margin: 0 0 0.8rem;
    letter-spacing: -0.02em;
}
.hero-title span {
    background: linear-gradient(90deg, #818cf8, #34d399);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.hero-sub {
    color: #8b8fa8;
    font-size: 1rem;
    font-weight: 300;
    max-width: 520px;
    line-height: 1.6;
    margin: 0;
}
.hero-badges {
    display: flex;
    gap: 10px;
    margin-top: 1.5rem;
    flex-wrap: wrap;
}
.hero-badge {
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 6px;
    padding: 5px 12px;
    font-size: 0.75rem;
    color: #9ca3af;
}
.hero-badge b { color: #e2e8f0; }

/* ── App grid ── */
.grid-section {
    padding: 2.5rem 3rem 3rem;
}
.section-label {
    font-family: 'Syne', sans-serif;
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: #4b5563;
    margin-bottom: 1.4rem;
}
.cards-grid {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 1.5rem;
}
@media (max-width: 900px) {
    .cards-grid { grid-template-columns: 1fr; }
    .hero-title { font-size: 2.2rem; }
    .grid-section { padding: 2rem 1.5rem; }
    .hero-wrap { padding: 2.5rem 1.5rem 2rem; }
}

/* ── App card ── */
.app-card {
    position: relative;
    background: #111118;
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 16px;
    padding: 0;
    overflow: hidden;
    cursor: pointer;
    transition: transform 0.25s ease, border-color 0.25s ease, box-shadow 0.25s ease;
    text-decoration: none;
    display: block;
}
.app-card:hover {
    transform: translateY(-4px);
    border-color: rgba(255,255,255,0.18);
    box-shadow: 0 20px 60px rgba(0,0,0,0.5);
}
.card-accent {
    height: 4px;
    width: 100%;
}
.card-body {
    padding: 1.6rem 1.8rem 1.4rem;
}
.card-icon-row {
    display: flex;
    align-items: center;
    gap: 1rem;
    margin-bottom: 1.1rem;
}
.card-icon {
    width: 48px; height: 48px;
    border-radius: 12px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1.4rem;
    flex-shrink: 0;
}
.card-tag {
    font-size: 0.68rem;
    font-weight: 600;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    padding: 3px 9px;
    border-radius: 4px;
}
.card-title {
    font-family: 'Syne', sans-serif;
    font-size: 1.25rem;
    font-weight: 700;
    color: #f0f0ff;
    margin: 0 0 0.5rem;
    letter-spacing: -0.01em;
}
.card-desc {
    font-size: 0.875rem;
    color: #6b7280;
    line-height: 1.6;
    margin: 0 0 1.3rem;
    font-weight: 300;
}
.card-features {
    list-style: none;
    padding: 0;
    margin: 0 0 1.4rem;
    display: flex;
    flex-direction: column;
    gap: 5px;
}
.card-features li {
    font-size: 0.8rem;
    color: #9ca3af;
    display: flex;
    align-items: center;
    gap: 7px;
}
.card-features li::before {
    content: '';
    width: 5px; height: 5px;
    border-radius: 50%;
    flex-shrink: 0;
}
.card-footer {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding-top: 1rem;
    border-top: 1px solid rgba(255,255,255,0.06);
}
.card-format {
    font-size: 0.72rem;
    color: #4b5563;
    font-weight: 500;
}
.launch-btn {
    font-size: 0.78rem;
    font-weight: 600;
    padding: 7px 18px;
    border-radius: 8px;
    border: none;
    cursor: pointer;
    letter-spacing: 0.02em;
    transition: opacity 0.15s;
    display: flex;
    align-items: center;
    gap: 6px;
}
.launch-btn:hover { opacity: 0.85; }

/* color themes per card */
.theme-indigo .card-accent { background: linear-gradient(90deg, #6366f1, #818cf8); }
.theme-indigo .card-icon   { background: rgba(99,102,241,0.15); }
.theme-indigo .card-tag    { background: rgba(99,102,241,0.15); color: #818cf8; }
.theme-indigo .card-features li::before { background: #6366f1; }
.theme-indigo .launch-btn  { background: #6366f1; color: white; }
.theme-indigo .app-card:hover { box-shadow: 0 20px 60px rgba(99,102,241,0.12); }

.theme-emerald .card-accent { background: linear-gradient(90deg, #10b981, #34d399); }
.theme-emerald .card-icon   { background: rgba(16,185,129,0.12); }
.theme-emerald .card-tag    { background: rgba(16,185,129,0.12); color: #34d399; }
.theme-emerald .card-features li::before { background: #10b981; }
.theme-emerald .launch-btn  { background: #10b981; color: white; }
.theme-emerald .app-card:hover { box-shadow: 0 20px 60px rgba(16,185,129,0.1); }

.theme-amber .card-accent { background: linear-gradient(90deg, #f59e0b, #fbbf24); }
.theme-amber .card-icon   { background: rgba(245,158,11,0.12); }
.theme-amber .card-tag    { background: rgba(245,158,11,0.12); color: #fbbf24; }
.theme-amber .card-features li::before { background: #f59e0b; }
.theme-amber .launch-btn  { background: #f59e0b; color: #0a0a0f; }
.theme-amber .app-card:hover { box-shadow: 0 20px 60px rgba(245,158,11,0.1); }

.theme-rose .card-accent { background: linear-gradient(90deg, #f43f5e, #fb7185); }
.theme-rose .card-icon   { background: rgba(244,63,94,0.12); }
.theme-rose .card-tag    { background: rgba(244,63,94,0.12); color: #fb7185; }
.theme-rose .card-features li::before { background: #f43f5e; }
.theme-rose .launch-btn  { background: #f43f5e; color: white; }
.theme-rose .app-card:hover { box-shadow: 0 20px 60px rgba(244,63,94,0.1); }

/* ── Active app view ── */
.nav-bar {
    background: #0d0d1a;
    border-bottom: 1px solid rgba(255,255,255,0.07);
    padding: 0.75rem 2rem;
    display: flex;
    align-items: center;
    gap: 1.5rem;
    position: sticky;
    top: 0;
    z-index: 100;
}
.nav-back {
    display: flex;
    align-items: center;
    gap: 7px;
    font-size: 0.82rem;
    color: #6b7280;
    cursor: pointer;
    background: none;
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 7px;
    padding: 6px 14px;
    font-family: 'DM Sans', sans-serif;
    transition: color 0.15s, border-color 0.15s;
}
.nav-back:hover { color: #e2e8f0; border-color: rgba(255,255,255,0.2); }
.nav-title {
    font-family: 'Syne', sans-serif;
    font-size: 0.95rem;
    font-weight: 700;
    color: #f0f0ff;
}
.nav-divider { color: #374151; }

/* ── Streamlit widget overrides inside app ── */
.stApp { color: #e2e8f0; }
</style>
""", unsafe_allow_html=True)

# ─── Session state ─────────────────────────────────────────────────────────────
if "active_app" not in st.session_state:
    st.session_state.active_app = None

APPS = [
    {
        "id": "tds",
        "theme": "indigo",
        "icon": "📄",
        "tag": "TDS",
        "title": "TDS Challan Extractor",
        "desc": "Parse ITNS 281 challan PDFs and export structured data to a formatted Excel workbook.",
        "features": ["Extracts TAN, CIN, BSR Code, Challan No", "Tax, Surcharge, Cess & Fee breakdowns", "Multi-file batch processing", "Auto-totals Excel export"],
        "format": "PDF → Excel",
        "module": "tds",
    },
    {
        "id": "epf",
        "theme": "emerald",
        "icon": "📋",
        "tag": "EPF",
        "title": "EPF Challan Consolidator",
        "desc": "Consolidate multiple EPFO Combined Challan PDFs into one detailed report with A/C breakdowns.",
        "features": ["A/C 01, 02, 10, 21 & 22 parsing", "Employer & Employee share split", "Grand Total + wages summary", "Per-establishment expanders"],
        "format": "PDF → Excel",
        "module": "epf",
    },
    {
        "id": "esic",
        "theme": "amber",
        "icon": "🏥",
        "tag": "ESIC",
        "title": "ESIC Challan Extractor",
        "desc": "Extract employer code, challan number, period and payment info from ESIC challan PDFs.",
        "features": ["Employer code & name extraction", "Challan period & transaction details", "Amount Paid with totals row", "Clean formatted Excel output"],
        "format": "PDF → Excel",
        "module": "esic",
    },
    {
        "id": "excel",
        "theme": "rose",
        "icon": "📊",
        "tag": "Excel",
        "title": "Excel File Consolidator",
        "desc": "Merge XLS, XLSX, XLSM and CSV files of any format into one clean spreadsheet with source tracking.",
        "features": ["Handles binary XLS & HTML-as-XLS", "Union or intersection column strategy", "Source File column auto-added", "File Summary sheet included"],
        "format": "XLS / XLSX / CSV → Excel",
        "module": "excel",
    },
]

# ─── HOME ─────────────────────────────────────────────────────────────────────
if st.session_state.active_app is None:

    st.markdown("""
    <div class="hero-wrap">
        <div class="hero-eyebrow">⚡ Compliance Automation Suite</div>
        <h1 class="hero-title">Your payroll compliance<br><span>tools, unified.</span></h1>
        <p class="hero-sub">Extract, consolidate and export TDS, EPF, ESIC challans and Excel files — all in one place. No switching tabs.</p>
        <div class="hero-badges">
            <span class="hero-badge"><b>4</b> tools</span>
            <span class="hero-badge"><b>PDF</b> extraction</span>
            <span class="hero-badge"><b>Excel</b> export</span>
            <span class="hero-badge"><b>Batch</b> processing</span>
        </div>
    </div>
    <div class="grid-section">
        <div class="section-label">Select a tool to get started</div>
        <div class="cards-grid">
    """, unsafe_allow_html=True)

    cols = st.columns(2)
    for i, app in enumerate(APPS):
        with cols[i % 2]:
            feats_html = "".join(f"<li>{f}</li>" for f in app["features"])
            st.markdown(f"""
            <div class="app-card theme-{app['theme']}">
                <div class="card-accent"></div>
                <div class="card-body">
                    <div class="card-icon-row">
                        <div class="card-icon">{app['icon']}</div>
                        <span class="card-tag">{app['tag']}</span>
                    </div>
                    <div class="card-title">{app['title']}</div>
                    <div class="card-desc">{app['desc']}</div>
                    <ul class="card-features">{feats_html}</ul>
                    <div class="card-footer">
                        <span class="card-format">{app['format']}</span>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button(f"Launch {app['title']} →", key=f"launch_{app['id']}", use_container_width=True):
                st.session_state.active_app = app["id"]
                st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)

# ─── ACTIVE APP ───────────────────────────────────────────────────────────────
else:
    active = next(a for a in APPS if a["id"] == st.session_state.active_app)

    # Nav bar
    col_back, col_title, col_spacer = st.columns([1, 4, 6])
    with col_back:
        if st.button("← Back to Suite", key="nav_back"):
            st.session_state.active_app = None
            # clear any sub-app session state to avoid bleed
            for key in list(st.session_state.keys()):
                if key not in ("active_app",):
                    del st.session_state[key]
            st.rerun()
    with col_title:
        st.markdown(
            f"<div style='padding-top:6px; font-family:Syne,sans-serif; font-size:0.95rem;"
            f"font-weight:700; color:#f0f0ff;'>{active['icon']} {active['title']}</div>",
            unsafe_allow_html=True
        )

    st.markdown("<hr style='border:none;border-top:1px solid rgba(255,255,255,0.07);margin:0.3rem 0 1rem'>", unsafe_allow_html=True)

    # ── Inject the selected app ─────────────────────────────────────────────
    if active["id"] == "tds":
        # ═══════════════════════════════════════════════════════════════════
        # TDS CHALLAN EXTRACTOR
        # ═══════════════════════════════════════════════════════════════════
        import pdfplumber
        import pandas as pd
        import re
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        st.markdown("""
        <style>
            .main-header { font-size: 28px; font-weight: 700; color: #1a1a2e; margin-bottom: 4px; }
            .sub-header { font-size: 14px; color: #6c757d; margin-bottom: 24px; }
            .metric-card { background: #f8f9fa; border-radius: 8px; padding: 16px; text-align: center; border: 1px solid #e9ecef; }
            .metric-value { font-size: 28px; font-weight: 700; color: #1a1a2e; }
            .metric-label { font-size: 12px; color: #6c757d; margin-top: 4px; }
            .stDataFrame { border-radius: 8px; }
            div[data-testid="stFileUploader"] { border: 2px dashed #dee2e6; border-radius: 12px; padding: 8px; }
        </style>
        """, unsafe_allow_html=True)

        st.markdown('<div class="main-header">📄 TDS Challan PDF Extractor</div>', unsafe_allow_html=True)
        st.markdown('<div class="sub-header">Upload ITNS 281 challan receipts — all data extracted into a single Excel sheet</div>', unsafe_allow_html=True)

        def tds_extract_value(text, label):
            patterns = [
                rf"{re.escape(label)}\s*[:\-]\s*(.+)",
                rf"{re.escape(label)}\s+(.+)",
            ]
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    return match.group(1).strip()
            return ""

        def tds_clean_amount(val):
            val = val.replace("₹", "").replace(",", "").strip()
            match = re.search(r"[\d]+(?:\.\d+)?", val)
            return float(match.group()) if match else 0.0

        def tds_extract_challan_data(pdf_file):
            with pdfplumber.open(pdf_file) as pdf:
                full_text = ""
                for page in pdf.pages:
                    full_text += page.extract_text() + "\n"
            data = {}
            fields = {
                "TAN": "TAN", "Name": "Name", "Assessment Year": "Assessment Year",
                "Financial Year": "Financial Year", "Major Head": "Major Head",
                "Minor Head": "Minor Head", "Nature of Payment": "Nature of Payment",
                "CIN": "CIN", "Mode of Payment": "Mode of Payment", "Bank Name": "Bank Name",
                "Bank Reference Number": "Bank Reference Number", "Date of Deposit": "Date of Deposit",
                "BSR code": "BSR Code", "Challan No": "Challan No", "Tender Date": "Tender Date",
            }
            for label, key in fields.items():
                data[key] = tds_extract_value(full_text, label)
            amount_match = re.search(r"Amount \(in Rs\.\)\s*[:\-]?\s*₹?\s*([\d,]+)", full_text)
            data["Amount (Rs.)"] = tds_clean_amount(amount_match.group(1)) if amount_match else 0.0
            amount_words_match = re.search(r"Amount \(in words\)\s*[:\-]?\s*(.+)", full_text)
            data["Amount (in words)"] = amount_words_match.group(1).strip() if amount_words_match else ""
            breakup_fields = {
                "Tax": r"A\s+Tax\s+₹?\s*([\d,]+)",
                "Surcharge": r"B\s+Surcharge\s+₹?\s*([\d,]+)",
                "Cess": r"C\s+Cess\s+₹?\s*([\d,]+)",
                "Interest": r"D\s+Interest\s+₹?\s*([\d,]+)",
                "Penalty": r"E\s+Penalty\s+₹?\s*([\d,]+)",
                "Fee u/s 234E": r"F\s+Fee under section 234E\s+₹?\s*([\d,]+)",
                "Total": r"Total \(A\+B\+C\+D\+E\+F\)\s+₹?\s*([\d,]+)",
            }
            for key, pattern in breakup_fields.items():
                match = re.search(pattern, full_text)
                data[key] = tds_clean_amount(match.group(1)) if match else 0.0
            itns_match = re.search(r"ITNS No\.\s*[:\-]?\s*(\d+)", full_text)
            data["ITNS No."] = itns_match.group(1).strip() if itns_match else ""
            return data

        def tds_create_excel(records):
            wb = Workbook()
            ws = wb.active
            ws.title = "TDS Challans"
            header_fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
            header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            sub_fill = PatternFill("solid", start_color="E8F4FD", end_color="E8F4FD")
            sub_font = Font(bold=True, name="Arial", size=9, color="1a1a2e")
            data_font = Font(name="Arial", size=9)
            alt_fill = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
            center = Alignment(horizontal="center", vertical="center")
            left = Alignment(horizontal="left", vertical="center")
            thin = Side(style="thin", color="DEE2E6")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.merge_cells("A1:T1")
            ws["A1"] = "TDS CHALLAN DETAILS — KAPSTON SERVICES LIMITED"
            ws["A1"].font = Font(bold=True, name="Arial", size=12, color="FFFFFF")
            ws["A1"].fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
            ws["A1"].alignment = center
            ws.row_dimensions[1].height = 28
            main_headers = [
                "S.No", "ITNS No.", "TAN", "Name", "Assessment Year", "Financial Year",
                "Nature of Payment", "CIN", "Mode of Payment", "Bank Name",
                "Bank Ref. No.", "Date of Deposit", "BSR Code", "Challan No", "Tender Date",
                "Tax (Rs.)", "Surcharge (Rs.)", "Cess (Rs.)", "Interest (Rs.)",
                "Penalty (Rs.)", "Fee u/s 234E (Rs.)", "Total Amount (Rs.)"
            ]
            ws.merge_cells("A2:A3"); ws.merge_cells("B2:B3")
            for col_idx, header in enumerate(main_headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = center; cell.border = border
                ws.cell(row=3, column=col_idx).font = sub_font
                ws.cell(row=3, column=col_idx).fill = sub_fill
                ws.cell(row=3, column=col_idx).alignment = center
                ws.cell(row=3, column=col_idx).border = border
            ws.row_dimensions[2].height = 20; ws.row_dimensions[3].height = 16
            for i, rec in enumerate(records):
                row = i + 4
                fill = alt_fill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
                values = [
                    i + 1, rec.get("ITNS No.", ""), rec.get("TAN", ""), rec.get("Name", ""),
                    rec.get("Assessment Year", ""), rec.get("Financial Year", ""),
                    rec.get("Nature of Payment", ""), rec.get("CIN", ""), rec.get("Mode of Payment", ""),
                    rec.get("Bank Name", ""), rec.get("Bank Reference Number", ""), rec.get("Date of Deposit", ""),
                    rec.get("BSR Code", ""), rec.get("Challan No", ""), rec.get("Tender Date", ""),
                    rec.get("Tax", 0), rec.get("Surcharge", 0), rec.get("Cess", 0),
                    rec.get("Interest", 0), rec.get("Penalty", 0), rec.get("Fee u/s 234E", 0), rec.get("Total", 0),
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = data_font; cell.fill = fill; cell.border = border
                    cell.alignment = center if col_idx == 1 else left
                    if col_idx >= 16: cell.number_format = '₹#,##0.00'
                ws.row_dimensions[row].height = 18
            total_row = len(records) + 4
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial", size=9)
            ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
            ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            ws.cell(row=total_row, column=1).alignment = center
            ws.merge_cells(f"A{total_row}:O{total_row}")
            for col_idx in range(16, 23):
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
                cell = ws.cell(row=total_row, column=col_idx, value=formula)
                cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
                cell.number_format = '₹#,##0.00'; cell.alignment = center; cell.border = border
            ws.row_dimensions[total_row].height = 20
            col_widths = [5,8,14,28,14,12,18,26,14,14,18,14,10,10,12,14,14,10,10,10,14,16]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.freeze_panes = "A4"
            output = BytesIO(); wb.save(output); output.seek(0)
            return output

        uploaded_files = st.file_uploader(
            "Upload challan PDF files", type=["pdf"], accept_multiple_files=True,
            help="Upload one or more ITNS 281 TDS challan PDF files", key="tds_uploader"
        )
        if uploaded_files:
            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(f'<div class="metric-card"><div class="metric-value">{len(uploaded_files)}</div><div class="metric-label">Files Uploaded</div></div>', unsafe_allow_html=True)
            records = []; errors = []
            with st.spinner("Extracting data from PDFs..."):
                for f in uploaded_files:
                    try:
                        data = tds_extract_challan_data(f)
                        data["_filename"] = f.name
                        records.append(data)
                    except Exception as e:
                        errors.append((f.name, str(e)))
            if errors:
                for fname, err in errors:
                    st.error(f"❌ {fname}: {err}")
            if records:
                total_amount = sum(r.get("Total", 0) for r in records)
                with col2:
                    st.markdown(f'<div class="metric-card"><div class="metric-value">{len(records)}</div><div class="metric-label">Extracted Successfully</div></div>', unsafe_allow_html=True)
                with col3:
                    st.markdown(f'<div class="metric-card"><div class="metric-value">₹{total_amount:,.0f}</div><div class="metric-label">Total TDS Amount</div></div>', unsafe_allow_html=True)
                st.markdown("### 📊 Extracted Data Preview")
                preview_cols = ["Nature of Payment","CIN","Challan No","Date of Deposit","BSR Code","Tax","Surcharge","Cess","Interest","Penalty","Fee u/s 234E","Total"]
                df = pd.DataFrame(records)
                df.insert(0, "S.No", range(1, len(df) + 1))
                df["File"] = df["_filename"]
                display_cols = ["S.No", "File"] + [c for c in preview_cols if c in df.columns]
                st.dataframe(df[display_cols], use_container_width=True, hide_index=True)
                st.markdown("### 💾 Export to Excel")
                excel_data = tds_create_excel(records)
                st.download_button(
                    label="⬇️ Download Excel File", data=excel_data,
                    file_name="TDS_Challans.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        else:
            st.info("👆 Upload one or more TDS challan PDF files to get started.")
            st.markdown("**Supported format:** ITNS 281 Challan Receipts from Income Tax Department")

    # ═══════════════════════════════════════════════════════════════════════
    elif active["id"] == "epf":
        # EPF CHALLAN CONSOLIDATOR
        # ═══════════════════════════════════════════════════════════════════
        import pdfplumber
        import re
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        def epf_extract_data(pdf_file):
            try:
                with pdfplumber.open(pdf_file) as pdf:
                    text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            except Exception as e:
                return None, f"Could not read PDF: {e}"
            if "EMPLOYEES' PROVIDENT FUND" not in text:
                return None, "Not an EPF challan — required header not found"
            data = {}
            def find(pattern, flags=0):
                m = re.search(pattern, text, flags)
                return m.group(1).strip() if m else ""
            data["TRRN"] = find(r"TRRN[:\s]*(\d+)")
            data["ECR Id"] = find(r"ECR\s*Id\s*(\d+)")
            data["LIN"] = find(r"LIN\s*[:\s]*(\d+)")
            m = re.search(r"Establishment Code\s*&\s*([A-Z0-9]+)\s+(.+?)\s+Dues for the wage month\s+(\w+)\s+(\d{4})", text)
            if m:
                data["Establishment Code"] = m.group(1)
                data["Company Name"] = m.group(2).strip()
                data["Wage Month"] = f"{m.group(3)} {m.group(4)}"
            else:
                data["Establishment Code"] = data["Company Name"] = data["Wage Month"] = ""
            m = re.search(r"Address\s*:\s*(.+?)(?=\nEPF|\nTotal)", text, re.DOTALL)
            data["Address"] = re.sub(r'\s+', ' ', m.group(1)).strip() if m else ""
            def parse_int_pair(pattern):
                m = re.search(pattern, text)
                if m:
                    return int(m.group(1).replace(",", "")), int(m.group(2).replace(",", ""))
                return "", ""
            a, b = parse_int_pair(r"Total Subscribers\s*:\s*([\d,]+)\s+([\d,]+)")
            data["Total Subscribers EPF"] = a; data["Total Subscribers EPS"] = b
            a, b = parse_int_pair(r"Total Wages\s*:\s*([\d,]+)\s+([\d,]+)")
            data["Total Wages EPF"] = a; data["Total Wages EPS"] = b
            def parse_row(label):
                m = re.search(label + r"\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)", text)
                return [int(g.replace(",", "")) for g in m.groups()] if m else [""] * 6
            for prefix, label in [("Admin", r"Administration Charges"), ("Employer", r"Employer.s Share Of"), ("Employee", r"Employee.s Share Of")]:
                vals = parse_row(label)
                for col, v in zip(["A/C.01","A/C.02","A/C.10","A/C.21","A/C.22","Total"], vals):
                    data[f"{prefix} {col}"] = v
            m = re.search(r"Grand Total\s*:\s*(.+?)\s+([\d,]+)\s*$", text, re.MULTILINE)
            if m:
                data["Grand Total (Words)"] = m.group(1).strip()
                data["Grand Total"] = int(m.group(2).replace(",", ""))
            else:
                data["Grand Total (Words)"] = data["Grand Total"] = ""
            m = re.search(r"Total remittance by Employer.*?([\d,]+)\s*$", text, re.MULTILINE)
            data["Total Remittance by Employer"] = int(m.group(1).replace(",", "")) if m else ""
            m = re.search(r"Total amount of uploaded ECR.*?([\d,]+)\s*$", text, re.MULTILINE)
            data["Total ECR Amount"] = int(m.group(1).replace(",", "")) if m else ""
            return data, None

        def epf_build_excel(records):
            wb = Workbook(); ws = wb.active; ws.title = "EPF Challan Summary"
            thin = Side(style="thin", color="BFBFBF")
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
            right_al = Alignment(horizontal="right", vertical="center")
            def mcell(row, col, value=None, bold=False, fg="000000", bg=None, align=None, fmt=None, size=9):
                c = ws.cell(row=row, column=col, value=value)
                c.font = Font(name="Arial", bold=bold, color=fg, size=size)
                if bg: c.fill = PatternFill("solid", start_color=bg)
                c.alignment = align or left_al; c.border = bdr
                if fmt: c.number_format = fmt
                return c
            INR_FMT = '#,##0;(#,##0);"-"'; INT_FMT = "#,##0"
            COLUMNS = [
                ("S.No",5,"S.No",False,True),("Source File",30,"Source File",False,False),
                ("Establishment\nCode",20,"Establishment Code",False,False),("Company Name",28,"Company Name",False,False),
                ("Address",42,"Address",False,False),("Wage\nMonth",13,"Wage Month",False,False),
                ("TRRN",18,"TRRN",False,False),("ECR Id",14,"ECR Id",False,False),("LIN",14,"LIN",False,False),
                ("EPF\nSubscribers",13,"Total Subscribers EPF",False,True),("EPS\nSubscribers",13,"Total Subscribers EPS",False,True),
                ("EPF Total\nWages (Rs.)",16,"Total Wages EPF",True,False),("EPS Total\nWages (Rs.)",16,"Total Wages EPS",True,False),
                ("Admin\nA/C.01",12,"Admin A/C.01",True,False),("Admin\nA/C.02",12,"Admin A/C.02",True,False),
                ("Admin\nA/C.10",12,"Admin A/C.10",True,False),("Admin\nA/C.21",12,"Admin A/C.21",True,False),
                ("Admin\nA/C.22",12,"Admin A/C.22",True,False),("Admin\nTotal (Rs.)",14,"Admin Total",True,False),
                ("Employer\nA/C.01",12,"Employer A/C.01",True,False),("Employer\nA/C.02",12,"Employer A/C.02",True,False),
                ("Employer\nA/C.10",12,"Employer A/C.10",True,False),("Employer\nA/C.21",12,"Employer A/C.21",True,False),
                ("Employer\nA/C.22",12,"Employer A/C.22",True,False),("Employer\nTotal (Rs.)",15,"Employer Total",True,False),
                ("Employee\nA/C.01",12,"Employee A/C.01",True,False),("Employee\nA/C.02",12,"Employee A/C.02",True,False),
                ("Employee\nA/C.10",12,"Employee A/C.10",True,False),("Employee\nA/C.21",12,"Employee A/C.21",True,False),
                ("Employee\nA/C.22",12,"Employee A/C.22",True,False),("Employee\nTotal (Rs.)",15,"Employee Total",True,False),
                ("Grand\nTotal (Rs.)",16,"Grand Total",True,False),("Grand Total (In Words)",45,"Grand Total (Words)",False,False),
                ("Total Remittance\nby Employer (Rs.)",18,"Total Remittance by Employer",True,False),
                ("Total ECR\nAmount (Rs.)",16,"Total ECR Amount",True,False),
            ]
            SECTIONS = [
                ("Establishment Information",1,9,"1F4E79"),("Subscribers & Wages",10,13,"1F4E79"),
                ("Administration Charges",14,19,"375623"),("Employer's Share",20,25,"843C0C"),
                ("Employee's Share",26,31,"7030A0"),("Totals",32,35,"1F4E79"),
            ]
            SEC_COLOR_MAP = {}
            for _, s, e, color in SECTIONS:
                for c in range(s, e+1): SEC_COLOR_MAP[c] = color
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
            c = ws.cell(row=1, column=1, value="EMPLOYEES' PROVIDENT FUND — CHALLAN CONSOLIDATED REPORT")
            c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="1F4E79")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 16
            for name, s, e, color in SECTIONS:
                ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
                c = ws.cell(row=2, column=s, value=name)
                c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", start_color=color)
                c.alignment = center; c.border = bdr
            ws.row_dimensions[3].height = 38
            for col_idx, (hdr, width, _, _, _) in enumerate(COLUMNS, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                bg_color = SEC_COLOR_MAP.get(col_idx, "1F4E79")
                c = ws.cell(row=3, column=col_idx, value=hdr)
                c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", start_color=bg_color)
                c.alignment = center; c.border = bdr
            ROW_BG = ["FFFFFF", "EBF3FB"]
            for row_idx, rec in enumerate(records, 1):
                excel_row = row_idx + 3; ws.row_dimensions[excel_row].height = 16
                bg = ROW_BG[row_idx % 2]
                for col_idx, (_, _, key, is_curr, is_int) in enumerate(COLUMNS, 1):
                    val = row_idx if key == "S.No" else rec.get(key, "")
                    fmt = INR_FMT if is_curr else (INT_FMT if is_int else None)
                    al = right_al if (is_curr or is_int) else left_al
                    mcell(excel_row, col_idx, val, fg="000000", bg=bg, align=al, fmt=fmt)
            total_row = len(records) + 4; ws.row_dimensions[total_row].height = 18
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
            c = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
            c.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
            c.fill = PatternFill("solid", start_color="FFF2CC"); c.alignment = center; c.border = bdr
            SUM_COLS = set(range(10, 32)) | {32, 34, 35}
            for col_idx, (_, _, _, is_curr, is_int) in enumerate(COLUMNS, 1):
                if col_idx not in SUM_COLS: continue
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}4:{col_letter}{total_row-1})"
                fmt = INR_FMT if is_curr else (INT_FMT if is_int else None)
                c = ws.cell(row=total_row, column=col_idx, value=formula)
                c.font = Font(name="Arial", bold=True, size=9)
                c.fill = PatternFill("solid", start_color="FFF2CC")
                c.alignment = right_al; c.border = bdr
                if fmt: c.number_format = fmt
            ws.freeze_panes = "A4"
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return buf

        st.title("📋 EPF Challan Consolidator")
        st.markdown("Upload any number of **EPF Combined Challan PDFs** to merge them into a single formatted Excel report. Non-EPF files are flagged automatically.")
        st.markdown("---")
        uploaded = st.file_uploader("Upload EPF Challan PDFs", type=["pdf"], accept_multiple_files=True,
            help="EPFO Combined Challan (A/C No. 01, 02, 10, 21 & 22)", key="epf_uploader")
        if uploaded:
            valid, results = [], []
            for f in uploaded:
                data, err = epf_extract_data(f)
                if err:
                    results.append({"file": f.name, "ok": False, "detail": err})
                else:
                    data["Source File"] = f.name
                    valid.append(data)
                    gt = data.get("Grand Total", "")
                    gt_str = f"₹{gt:,}" if isinstance(gt, int) else str(gt)
                    detail = (f"Estab: **{data.get('Establishment Code','')}** | Month: **{data.get('Wage Month','')}** | Grand Total: **{gt_str}**")
                    results.append({"file": f.name, "ok": True, "detail": detail})
            st.markdown("### 📂 File Processing Results")
            for r in results:
                icon = "✅" if r["ok"] else "❌"; label = "Valid EPF Challan" if r["ok"] else "Wrong Format"
                c1, c2, c3 = st.columns([3, 2, 6])
                c1.write(f"`{r['file']}`"); c2.write(f"{icon} {label}"); c3.markdown(r["detail"])
            st.markdown("---")
            if valid:
                st.success(f"✅ **{len(valid)} valid challan(s)** ready to consolidate.")
                if st.button("📥 Generate Consolidated Excel", type="primary", use_container_width=True, key="epf_generate"):
                    with st.spinner("Building Excel report..."):
                        excel_buf = epf_build_excel(valid)
                    st.download_button(label="⬇️ Download EPF_Challan_Consolidated.xlsx", data=excel_buf,
                        file_name="EPF_Challan_Consolidated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    st.markdown("### 📊 Summary")
                    total_subs = sum(r.get("Total Subscribers EPF", 0) or 0 for r in valid)
                    total_wages = sum(r.get("Total Wages EPF", 0) or 0 for r in valid)
                    total_emp = sum(r.get("Employee Total", 0) or 0 for r in valid)
                    total_empr = sum(r.get("Employer Total", 0) or 0 for r in valid)
                    total_grand = sum(r.get("Grand Total", 0) or 0 for r in valid)
                    k1,k2,k3,k4,k5 = st.columns(5)
                    k1.metric("Establishments", len(valid)); k2.metric("Total EPF Members", f"{total_subs:,}")
                    k3.metric("Total Wages", f"Rs.{total_wages:,.0f}"); k4.metric("Employee Share", f"Rs.{total_emp:,.0f}")
                    k5.metric("Grand Total", f"Rs.{total_grand:,.0f}")
                    st.markdown("### 🏢 Per-Establishment Details")
                    for i, r in enumerate(valid, 1):
                        gt = r.get("Grand Total", 0) or 0
                        with st.expander(f"{i}. {r.get('Establishment Code','')} — {r.get('Company Name','')} | {r.get('Wage Month','')} | Rs.{gt:,}"):
                            col_a, col_b, col_c = st.columns(3)
                            col_a.markdown(f"**Address:** {r.get('Address','')}"); col_a.markdown(f"**TRRN:** `{r.get('TRRN','')}`"); col_a.markdown(f"**LIN:** `{r.get('LIN','')}`")
                            col_b.metric("EPF Subscribers", r.get("Total Subscribers EPF","")); col_b.metric("EPS Subscribers", r.get("Total Subscribers EPS","")); col_b.metric("Total Wages", f"Rs.{r.get('Total Wages EPF',0):,}")
                            col_c.metric("Admin Charges", f"Rs.{r.get('Admin Total',0):,}"); col_c.metric("Employer Share", f"Rs.{r.get('Employer Total',0):,}"); col_c.metric("Employee Share", f"Rs.{r.get('Employee Total',0):,}")
            else:
                st.error("No valid EPF challans found. Please upload correct EPFO Combined Challan PDFs.")
        else:
            st.info("👆 Upload one or more EPF Challan PDFs above to get started.")
            with st.expander("ℹ️ How it works"):
                st.markdown("""
                1. **Upload** — Drop any number of EPFO Combined Challan PDFs
                2. **Parse** — Each PDF is automatically validated and parsed
                3. **Export** — Click Generate to download a consolidated Excel
                4. **Wrong Format** — Any non-EPF PDFs are flagged and skipped
                """)

    # ═══════════════════════════════════════════════════════════════════════
    elif active["id"] == "esic":
        # ESIC CHALLAN EXTRACTOR
        # ═══════════════════════════════════════════════════════════════════
        import pdfplumber
        import pandas as pd
        import re
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        st.markdown("""
            <style>
            @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
            .esic-title { font-family:'IBM Plex Mono',monospace !important; font-size:1.6rem !important; color:#1a1a2e !important; border-bottom:3px solid #c0392b; padding-bottom:0.4rem; }
            </style>
        """, unsafe_allow_html=True)

        st.markdown('<h1 class="esic-title">ESIC Challan Extractor</h1>', unsafe_allow_html=True)
        st.markdown("Upload one or more ESIC Challan PDFs to extract and export data to Excel.")

        ESIC_FIELDS = {
            "Employer's Code No": "Employer Code No", "Employer's Name": "Employer Name",
            "Challan Period": "Challan Period", "Challan Number": "Challan Number",
            "Challan Created Date": "Challan Created Date", "Challan Submitted Date": "Challan Submitted Date",
            "Amount Paid": "Amount Paid", "Transaction Number": "Transaction Number",
            "Transaction status": "Transaction Status",
        }

        def esic_extract_from_pdf(file_bytes):
            data = {}
            with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            for field, col in ESIC_FIELDS.items():
                pattern = re.escape(field) + r"[\s:]*([^\n]+)"
                match = re.search(pattern, text, re.IGNORECASE)
                data[col] = match.group(1).strip().rstrip("*").strip() if match else ""
            return data

        def esic_create_excel(records):
            wb = Workbook(); ws = wb.active; ws.title = "ESIC Challans"
            headers = ["Source File"] + list(ESIC_FIELDS.values())
            header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", start_color="1A1A2E")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            data_font = Font(name="Arial", size=10)
            alt_fill = PatternFill("solid", start_color="F2F2EF")
            center_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = border
            ws.row_dimensions[1].height = 30
            for row_idx, record in enumerate(records, 2):
                fill = PatternFill("solid", start_color="FFFFFF") if row_idx % 2 == 0 else alt_fill
                for col_idx, header in enumerate(headers, 1):
                    val = record.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = data_font; cell.fill = fill; cell.alignment = center_align; cell.border = border
                ws.row_dimensions[row_idx].height = 20
            col_widths = [30,22,28,14,22,22,22,14,22,26]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            total_row = len(records) + 2
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
            ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="C0392B")
            ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF")
            ws.cell(row=total_row, column=1).alignment = center_align
            amt_col = list(ESIC_FIELDS.values()).index("Amount Paid") + 2
            total_formula = f"=SUM({get_column_letter(amt_col)}2:{get_column_letter(amt_col)}{total_row-1})"
            total_cell = ws.cell(row=total_row, column=amt_col, value=total_formula)
            total_cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            total_cell.fill = PatternFill("solid", start_color="C0392B")
            total_cell.alignment = center_align; total_cell.border = border
            ws.freeze_panes = "A2"
            output = BytesIO(); wb.save(output); output.seek(0)
            return output

        uploaded_files = st.file_uploader("Upload ESIC Challan PDFs", type=["pdf"], accept_multiple_files=True,
            help="You can upload multiple PDFs at once", key="esic_uploader")
        if uploaded_files:
            st.markdown(f"**{len(uploaded_files)} file(s) uploaded**")
            records = []; errors = []
            for f in uploaded_files:
                try:
                    record = esic_extract_from_pdf(f.read())
                    record["_filename"] = f.name
                    records.append(record)
                except Exception as e:
                    errors.append(f"{f.name}: {e}")
            if errors:
                for err in errors: st.error(f"⚠️ {err}")
            if records:
                display_records = [{"Source File": r["_filename"], **{k: v for k, v in r.items() if k != "_filename"}} for r in records]
                df = pd.DataFrame(display_records)
                st.markdown("### Preview")
                st.dataframe(df, use_container_width=True)
                try:
                    total_amt = sum(float(r.get('Amount Paid', 0) or 0) for r in display_records)
                    st.markdown(f"**{len(records)} record(s) extracted** | Total Amount: ₹{total_amt:,.2f}")
                except Exception:
                    st.markdown(f"**{len(records)} record(s) extracted**")
                if st.button("⬇ Download Excel", key="esic_dl"):
                    excel_file = esic_create_excel(display_records)
                    st.download_button(label="📥 Click to Save Excel File", data=excel_file,
                        file_name="ESIC_Challans.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Upload PDFs above to get started.")

    # ═══════════════════════════════════════════════════════════════════════
    elif active["id"] == "excel":
        # EXCEL FILE CONSOLIDATOR
        # ═══════════════════════════════════════════════════════════════════
        import pandas as pd
        import numpy as np
        from io import BytesIO
        from pathlib import Path
        from bs4 import BeautifulSoup
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import xlrd
        import warnings
        warnings.filterwarnings("ignore")

        st.markdown("""
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
            .stat-card { background:white; border-radius:12px; padding:1.2rem 1.5rem; box-shadow:0 2px 12px rgba(0,0,0,0.07); border-left:4px solid #2d6a9f; margin-bottom:0.5rem; }
            .stat-card .label { font-size:0.78rem; color:#64748b; font-weight:500; text-transform:uppercase; letter-spacing:0.05em; }
            .stat-card .value { font-size:1.6rem; font-weight:700; color:#1e3a5f; margin-top:0.1rem; }
            .stat-card .sub { font-size:0.82rem; color:#94a3b8; }
            .file-badge { display:inline-block; padding:2px 10px; border-radius:20px; font-size:0.75rem; font-weight:600; margin:2px; }
            .badge-html { background:#dbeafe; color:#1d4ed8; }
            .badge-xls  { background:#dcfce7; color:#166534; }
            .badge-xlsx { background:#fef9c3; color:#854d0e; }
            .badge-csv  { background:#fce7f3; color:#9d174d; }
            .step-box { background:white; border-radius:10px; padding:1rem 1.2rem; margin-bottom:0.8rem; box-shadow:0 1px 6px rgba(0,0,0,0.06); border:1px solid #e2e8f0; }
            .step-num { display:inline-block; width:26px; height:26px; background:#2d6a9f; color:white; border-radius:50%; text-align:center; line-height:26px; font-size:0.8rem; font-weight:700; margin-right:8px; }
            .error-box { background:#fef2f2; border:1.5px solid #fca5a5; border-radius:10px; padding:0.8rem 1.2rem; margin:0.5rem 0; }
        </style>
        """, unsafe_allow_html=True)

        def excel_detect_format(file_bytes, filename):
            ext = Path(filename).suffix.lower()
            if ext == ".csv": return "csv"
            if ext in (".xlsx", ".xlsm"): return "xlsx"
            if file_bytes[:4] == b"\xd0\xcf\x11\xe0": return "xls_real"
            snippet = file_bytes[:200].decode("utf-8", errors="ignore").lower()
            if "<table" in snippet or "<html" in snippet or "<style" in snippet: return "xls_html"
            return "xls_real"

        def excel_read_html_xls(file_bytes, source_name):
            content = file_bytes.decode("utf-8", errors="ignore")
            soup = BeautifulSoup(content, "html.parser")
            table = soup.find("table")
            if not table: raise ValueError("No HTML table found in file.")
            header_row = table.find("tr")
            headers = [th.get_text(strip=True) for th in header_row.find_all(["th", "td"])]
            rows = []
            for tr in table.find_all("tr")[1:]:
                cells = [td.get_text(separator=" ", strip=True) for td in tr.find_all("td")]
                if any(c.strip() for c in cells): rows.append(cells)
            if not rows: raise ValueError("Table found but no data rows.")
            ncols = max(len(r) for r in rows)
            headers = (headers + [""] * ncols)[:ncols]
            return pd.DataFrame(rows, columns=headers), "HTML-as-XLS"

        def excel_read_real_xls(file_bytes, source_name, sheet_choice):
            bio = BytesIO(file_bytes)
            sheets = pd.read_excel(bio, engine="xlrd", sheet_name=None, dtype=str)
            if not sheets: raise ValueError("No sheets found.")
            df = sheets[sheet_choice] if (sheet_choice and sheet_choice in sheets) else list(sheets.values())[0]
            return df.dropna(how="all").reset_index(drop=True), "XLS (Binary)"

        def excel_read_xlsx(file_bytes, source_name, sheet_choice):
            bio = BytesIO(file_bytes)
            sheets = pd.read_excel(bio, engine="openpyxl", sheet_name=None, dtype=str)
            if not sheets: raise ValueError("No sheets found.")
            df = sheets[sheet_choice] if (sheet_choice and sheet_choice in sheets) else list(sheets.values())[0]
            return df.dropna(how="all").reset_index(drop=True), "XLSX"

        def excel_read_csv(file_bytes, source_name, encoding="utf-8"):
            for enc in [encoding, "latin-1", "cp1252"]:
                try:
                    df = pd.read_csv(BytesIO(file_bytes), dtype=str, encoding=enc)
                    return df.dropna(how="all").reset_index(drop=True), "CSV"
                except Exception: continue
            raise ValueError("Could not decode CSV file.")

        def excel_get_sheet_names(file_bytes, filename):
            ext = Path(filename).suffix.lower()
            try:
                if ext in (".xlsx", ".xlsm"):
                    return pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl").sheet_names
                elif ext in (".xls",):
                    fmt = excel_detect_format(file_bytes, filename)
                    if fmt == "xls_real":
                        return pd.ExcelFile(BytesIO(file_bytes), engine="xlrd").sheet_names
            except Exception: pass
            return []

        def excel_parse_file(file_bytes, filename, sheet_choice=None):
            fmt = excel_detect_format(file_bytes, filename)
            if fmt == "csv": return excel_read_csv(file_bytes, filename)
            elif fmt == "xlsx": return excel_read_xlsx(file_bytes, filename, sheet_choice)
            elif fmt == "xls_html": return excel_read_html_xls(file_bytes, filename)
            else: return excel_read_real_xls(file_bytes, filename, sheet_choice)

        def excel_build_output(combined, include_summary):
            wb = Workbook()
            H_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
            H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            D_FONT = Font(name="Arial", size=9)
            ALT_FILL = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
            SRC_FILL = PatternFill("solid", start_color="FFF9C4", end_color="FFF9C4")
            SRC_FONT = Font(name="Arial", size=9, color="7B4F00")
            CENTER = Alignment(horizontal="center", vertical="center")
            LEFT = Alignment(horizontal="left", vertical="center")
            thin = Side(style="thin", color="CCCCCC")
            BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws = wb.active; ws.title = "Consolidated Data"
            cols = list(combined.columns)
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CENTER; cell.border = BORDER
                if col == "Source File":
                    cell.fill = PatternFill("solid", start_color="B8860B", end_color="B8860B")
            for ri, row_data in enumerate(combined.itertuples(index=False), 2):
                alt = ri % 2 == 0
                for ci, val in enumerate(row_data, 1):
                    v = "" if (val is None or (isinstance(val, float) and np.isnan(val))) else str(val)
                    cell = ws.cell(row=ri, column=ci, value=v)
                    cell.border = BORDER; cell.font = D_FONT
                    col_name = cols[ci - 1]
                    if col_name == "Source File":
                        cell.fill = SRC_FILL; cell.font = SRC_FONT; cell.alignment = CENTER
                    else:
                        cell.fill = ALT_FILL if alt else PatternFill()
                        cell.alignment = CENTER if ci <= 3 else LEFT
            for ci, col in enumerate(cols, 1):
                max_len = len(str(col))
                for ri in range(2, min(len(combined) + 2, 200)):
                    v = ws.cell(row=ri, column=ci).value or ""
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 45)
            ws.row_dimensions[1].height = 22; ws.freeze_panes = "B2"
            if include_summary and "Source File" in combined.columns:
                ws2 = wb.create_sheet("File Summary")
                for ci, h in enumerate(["Source File", "Rows", "Columns"], 1):
                    cell = ws2.cell(row=1, column=ci, value=h)
                    cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CENTER; cell.border = BORDER
                groups = combined.groupby("Source File", sort=False)
                for ri, (src, grp) in enumerate(groups, 2):
                    alt = ri % 2 == 0
                    for ci, val in enumerate([src, len(grp), len(combined.columns) - 1], 1):
                        cell = ws2.cell(row=ri, column=ci, value=val)
                        cell.font = D_FONT; cell.alignment = CENTER; cell.border = BORDER
                        if alt: cell.fill = ALT_FILL
                tr = len(groups) + 2
                for ci, val in enumerate(["TOTAL", len(combined), ""], 1):
                    cell = ws2.cell(row=tr, column=ci, value=val)
                    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                    cell.fill = H_FILL; cell.alignment = CENTER; cell.border = BORDER
                for ci, w in enumerate([40, 12, 12], 1):
                    ws2.column_dimensions[get_column_letter(ci)].width = w
                ws2.row_dimensions[1].height = 22
            buf = BytesIO(); wb.save(buf)
            return buf.getvalue()

        # Session state for excel app
        if "xl_parsed_files" not in st.session_state:
            st.session_state.xl_parsed_files = {}
            st.session_state.xl_file_formats = {}
            st.session_state.xl_file_errors = {}
            st.session_state.xl_combined_df = None

        # Sidebar options
        with st.sidebar:
            st.markdown("## ⚙️ Options"); st.markdown("---")
            output_filename = st.text_input("Output filename", value="Consolidated_Output")
            include_summary = st.checkbox("Add File Summary sheet", value=True)
            st.markdown("### 🧹 Column Handling")
            col_strategy = st.selectbox("Mismatched columns", ["Union (keep all columns)", "Intersection (common columns only)"])
            st.markdown("### 🔢 Row Numbering")
            add_row_num = st.checkbox("Add global row number column", value=False)
            st.markdown("---")
            st.markdown("### 📋 Supported Formats")
            for fmt, badge in [("`.xls` — Binary XLS","badge-xls"),("`.xls` — HTML-as-XLS","badge-html"),("`.xlsx` / `.xlsm`","badge-xlsx"),("`.csv`","badge-csv")]:
                st.markdown(f'<span class="file-badge {badge}">{fmt}</span>', unsafe_allow_html=True)

        st.markdown("""
        <div style="background:linear-gradient(135deg,#1e3a5f 0%,#2d6a9f 50%,#1a8cff 100%);border-radius:16px;padding:2.5rem 2rem;margin-bottom:1.5rem;color:white;box-shadow:0 8px 32px rgba(30,58,95,0.18);">
            <h1 style="font-size:2.1rem;font-weight:700;margin:0 0 0.3rem;letter-spacing:-0.5px;">📊 Excel File Consolidator</h1>
            <p style="font-size:1rem;opacity:0.88;margin:0;">Merge multiple Excel / CSV files of any format into one clean, formatted spreadsheet.</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="step-box"><span class="step-num">1</span> <b>Upload your files</b></div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("Drop files here (XLS, XLSX, CSV — any mix)", type=["xls","xlsx","xlsm","csv"],
            accept_multiple_files=True, label_visibility="collapsed", key="xl_uploader")

        if uploaded:
            st.markdown(f"**{len(uploaded)} file(s) selected**")
            sheet_selections = {}
            for uf in uploaded:
                file_bytes = uf.read(); uf.seek(0)
                sheets = excel_get_sheet_names(file_bytes, uf.name)
                if len(sheets) > 1:
                    with st.expander(f"📋 `{uf.name}` has {len(sheets)} sheets — choose one"):
                        chosen = st.selectbox(f"Sheet for {uf.name}", sheets, key=f"xl_sheet_{uf.name}", label_visibility="collapsed")
                        sheet_selections[uf.name] = chosen

            st.markdown('<div class="step-box"><span class="step-num">2</span> <b>Parse & preview files</b></div>', unsafe_allow_html=True)
            if st.button("🔍 Parse All Files", key="xl_parse"):
                st.session_state.xl_parsed_files = {}; st.session_state.xl_file_formats = {}
                st.session_state.xl_file_errors = {}; st.session_state.xl_combined_df = None
                progress = st.progress(0); status = st.empty()
                for i, uf in enumerate(uploaded):
                    status.text(f"Parsing {uf.name}…")
                    try:
                        file_bytes = uf.read(); sheet_ch = sheet_selections.get(uf.name)
                        df, fmt = excel_parse_file(file_bytes, uf.name, sheet_ch)
                        df = df.dropna(how="all").reset_index(drop=True)
                        st.session_state.xl_parsed_files[uf.name] = df
                        st.session_state.xl_file_formats[uf.name] = fmt
                    except Exception as e:
                        st.session_state.xl_file_errors[uf.name] = str(e)
                    progress.progress((i + 1) / len(uploaded))
                status.empty(); progress.empty()

            if st.session_state.xl_parsed_files:
                for fname, df in st.session_state.xl_parsed_files.items():
                    fmt = st.session_state.xl_file_formats.get(fname, "")
                    badge_cls = {"HTML-as-XLS":"badge-html","XLS (Binary)":"badge-xls","XLSX":"badge-xlsx","CSV":"badge-csv"}.get(fmt,"badge-xls")
                    with st.expander(f"✅  `{fname}` — {len(df):,} rows × {len(df.columns)} cols"):
                        st.markdown(f'<span class="file-badge {badge_cls}">{fmt}</span>', unsafe_allow_html=True)
                        st.dataframe(df.head(5), use_container_width=True, hide_index=True)

            for fname, err in st.session_state.xl_file_errors.items():
                st.markdown(f'<div class="error-box">❌ <b>{fname}</b>: {err}</div>', unsafe_allow_html=True)

            if st.session_state.xl_parsed_files:
                st.markdown('<div class="step-box"><span class="step-num">3</span> <b>Consolidate</b></div>', unsafe_allow_html=True)
                if st.button("⚡ Consolidate All Files", key="xl_consolidate"):
                    dfs = []
                    for fname, df in st.session_state.xl_parsed_files.items():
                        df = df.copy(); df.insert(0, "Source File", fname); dfs.append(df)
                    if col_strategy.startswith("Union"):
                        combined = pd.concat(dfs, ignore_index=True, sort=False)
                    else:
                        common = set(dfs[0].columns)
                        for d in dfs[1:]: common &= set(d.columns)
                        common = sorted(common, key=lambda c: list(dfs[0].columns).index(c) if c in dfs[0].columns else 999)
                        combined = pd.concat([d[list(common)] for d in dfs], ignore_index=True, sort=False)
                    if add_row_num: combined.insert(1, "Row No.", range(1, len(combined) + 1))
                    st.session_state.xl_combined_df = combined

                if st.session_state.xl_combined_df is not None:
                    combined = st.session_state.xl_combined_df
                    file_count = combined["Source File"].nunique() if "Source File" in combined.columns else len(st.session_state.xl_parsed_files)
                    c1, c2, c3, c4 = st.columns(4)
                    for col_obj, label, value, sub in [
                        (c1,"Total Rows",f"{len(combined):,}","across all files"),
                        (c2,"Total Columns",f"{len(combined.columns):,}","in output sheet"),
                        (c3,"Files Merged",f"{file_count}","source files"),
                        (c4,"Output Size",f"~{len(combined)*len(combined.columns)//1000}K","cells"),
                    ]:
                        col_obj.markdown(f'<div class="stat-card"><div class="label">{label}</div><div class="value">{value}</div><div class="sub">{sub}</div></div>', unsafe_allow_html=True)
                    st.markdown("**Preview — first 50 rows**")
                    st.dataframe(combined.head(50), use_container_width=True, hide_index=True)
                    st.markdown('<div class="step-box"><span class="step-num">4</span> <b>Download</b></div>', unsafe_allow_html=True)
                    dl_col1, dl_col2 = st.columns(2)
                    with dl_col1:
                        with st.spinner("Building Excel file…"):
                            xlsx_bytes = excel_build_output(combined, include_summary)
                        st.download_button(label="⬇️ Download as Excel (.xlsx)", data=xlsx_bytes,
                            file_name=f"{output_filename}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    with dl_col2:
                        csv_bytes = combined.to_csv(index=False).encode("utf-8-sig")
                        st.download_button(label="⬇️ Download as CSV", data=csv_bytes,
                            file_name=f"{output_filename}.csv", mime="text/csv")
        else:
            st.markdown("""
            <div style="text-align:center;padding:3rem 1rem;color:#94a3b8;">
                <div style="font-size:3.5rem;margin-bottom:1rem;">📂</div>
                <div style="font-size:1.1rem;font-weight:500;">Upload your Excel or CSV files above to get started</div>
                <div style="font-size:0.9rem;margin-top:0.5rem;">Supports XLS (binary & HTML), XLSX, XLSM, and CSV</div>
            </div>""", unsafe_allow_html=True)
        st.markdown("---")
        st.markdown('<p style="text-align:center;color:#94a3b8;font-size:0.8rem;">Excel Consolidator · Handles XLS (binary & HTML-as-XLS), XLSX, XLSM, CSV · Source column auto-added</p>', unsafe_allow_html=True)
