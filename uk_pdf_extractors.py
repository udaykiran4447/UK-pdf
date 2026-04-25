import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="UK",
    page_icon="U",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── Session state ──
if "color_mode" not in st.session_state:
    st.session_state.color_mode = "System"
if "active_app" not in st.session_state:
    st.session_state.active_app = None
if "search_query" not in st.session_state:
    st.session_state.search_query = ""

APPS = {
    "ESIC Challan Extractor": "Extract ESIC Challan PDF data into Excel",
    "TDS Challan Extractor":  "Extract ITNS 281 TDS Challan PDF data into Excel",
}

mode = st.session_state.color_mode

# ── Color tokens per mode ──
DARK = {
    "bg":        "#0a0a0a",
    "bg2":       "#111111",
    "surface":   "#161616",
    "surface2":  "#1e1e1e",
    "border":    "#272727",
    "border2":   "#333333",
    "text":      "#f0f0f0",
    "text2":     "#888888",
    "text3":     "#555555",
    "accent":    "#f5f500",
    "accent_fg": "#0a0a0a",
    "btn_bg":    "#f0f0f0",
    "btn_text":  "#0a0a0a",
    "tag_bg":    "#1e1e1e",
    "tag_text":  "#aaaaaa",
    "hero_text": "#f0f0f0",
    "nav_bg":    "#0a0a0a",
}
LIGHT = {
    "bg":        "#f8f8f5",
    "bg2":       "#f0f0ec",
    "surface":   "#ffffff",
    "surface2":  "#f4f4f0",
    "border":    "#e0e0d8",
    "border2":   "#cccccc",
    "text":      "#111111",
    "text2":     "#666666",
    "text3":     "#aaaaaa",
    "accent":    "#1a1a1a",
    "accent_fg": "#ffffff",
    "btn_bg":    "#111111",
    "btn_text":  "#ffffff",
    "tag_bg":    "#eeeeea",
    "tag_text":  "#555555",
    "hero_text": "#111111",
    "nav_bg":    "#f8f8f5",
}

import_prefers = """
@media (prefers-color-scheme: dark) {
  :root { --bg:#0a0a0a; --bg2:#111111; --surface:#161616; --surface2:#1e1e1e;
    --border:#272727; --border2:#333333; --text:#f0f0f0; --text2:#888888;
    --text3:#555555; --accent:#f5f500; --accent-fg:#0a0a0a;
    --btn-bg:#f0f0f0; --btn-text:#0a0a0a; --tag-bg:#1e1e1e; --tag-text:#aaaaaa;
    --hero-text:#f0f0f0; --nav-bg:#0a0a0a; }
}
@media (prefers-color-scheme: light) {
  :root { --bg:#f8f8f5; --bg2:#f0f0ec; --surface:#ffffff; --surface2:#f4f4f0;
    --border:#e0e0d8; --border2:#cccccc; --text:#111111; --text2:#666666;
    --text3:#aaaaaa; --accent:#1a1a1a; --accent-fg:#ffffff;
    --btn-bg:#111111; --btn-text:#ffffff; --tag-bg:#eeeeea; --tag-text:#555555;
    --hero-text:#111111; --nav-bg:#f8f8f5; }
}
"""

if mode == "Dark":
    T = DARK
    css_vars = f"""
    :root {{
      --bg:{T['bg']}; --bg2:{T['bg2']}; --surface:{T['surface']}; --surface2:{T['surface2']};
      --border:{T['border']}; --border2:{T['border2']}; --text:{T['text']}; --text2:{T['text2']};
      --text3:{T['text3']}; --accent:{T['accent']}; --accent-fg:{T['accent_fg']};
      --btn-bg:{T['btn_bg']}; --btn-text:{T['btn_text']}; --tag-bg:{T['tag_bg']};
      --tag-text:{T['tag_text']}; --hero-text:{T['hero_text']}; --nav-bg:{T['nav_bg']};
    }}"""
elif mode == "Light":
    T = LIGHT
    css_vars = f"""
    :root {{
      --bg:{T['bg']}; --bg2:{T['bg2']}; --surface:{T['surface']}; --surface2:{T['surface2']};
      --border:{T['border']}; --border2:{T['border2']}; --text:{T['text']}; --text2:{T['text2']};
      --text3:{T['text3']}; --accent:{T['accent']}; --accent-fg:{T['accent_fg']};
      --btn-bg:{T['btn_bg']}; --btn-text:{T['btn_text']}; --tag-bg:{T['tag_bg']};
      --tag-text:{T['tag_text']}; --hero-text:{T['hero_text']}; --nav-bg:{T['nav_bg']};
    }}"""
    T = LIGHT
else:
    T = DARK  # fallback for Streamlit rendering
    css_vars = import_prefers

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Instrument+Sans:wght@400;500;600&family=JetBrains+Mono:wght@300;400;500&display=swap');

{css_vars}

*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

html, body, [class*="css"] {{
    font-family: 'Instrument Sans', sans-serif !important;
    background: var(--bg) !important;
    color: var(--text) !important;
}}
.stApp {{ background: var(--bg) !important; }}
.main .block-container {{
    padding: 0 !important;
    max-width: 100% !important;
}}

/* ═══ TICKER BAR ═══ */
.ticker-wrap {{
    background: var(--accent);
    overflow: hidden;
    white-space: nowrap;
    padding: 0.38rem 0;
    border-bottom: 1px solid var(--border);
}}
.ticker-inner {{
    display: inline-block;
    animation: ticker 28s linear infinite;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.68rem;
    font-weight: 500;
    color: var(--accent-fg);
    letter-spacing: 0.1em;
}}
@keyframes ticker {{
    0%   {{ transform: translateX(100vw); }}
    100% {{ transform: translateX(-100%); }}
}}

/* ═══ NAVBAR ═══ */
.navbar {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0 2.5rem;
    height: 60px;
    background: var(--nav-bg);
    border-bottom: 1px solid var(--border);
    position: sticky;
    top: 0;
    z-index: 1000;
}}
.nav-logo {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: 1.6rem;
    color: var(--text) !important;
    letter-spacing: 0.08em;
    line-height: 1;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}}
.nav-logo-badge {{
    width: 28px; height: 28px;
    background: var(--text);
    color: var(--bg);
    border-radius: 4px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-family: 'Bebas Neue', sans-serif;
    font-size: 1rem;
    line-height: 1;
}}
.nav-links {{
    display: flex;
    align-items: center;
    gap: 0.2rem;
    font-size: 0.78rem;
    font-weight: 500;
    letter-spacing: 0.04em;
}}
.nav-pill {{
    padding: 0.35rem 0.9rem;
    border-radius: 999px;
    color: var(--text2);
    cursor: default;
    transition: background 0.15s, color 0.15s;
    white-space: nowrap;
}}
.nav-pill:hover {{ background: var(--surface2); color: var(--text); }}
.nav-pill.active {{ background: var(--surface2); color: var(--text); }}

/* ═══ HERO ═══ */
.hero-section {{
    padding: 5rem 2.5rem 3.5rem;
    border-bottom: 1px solid var(--border);
    background: var(--bg);
    position: relative;
    overflow: hidden;
}}
.hero-eyebrow {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.65rem;
    letter-spacing: 0.3em;
    text-transform: uppercase;
    color: var(--text3);
    margin-bottom: 1.2rem;
}}
.hero-title {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: clamp(5rem, 14vw, 11rem);
    line-height: 0.9;
    color: var(--hero-text) !important;
    letter-spacing: -0.01em;
    margin-bottom: 1.8rem;
}}
.hero-title span {{
    color: var(--accent);
    -webkit-text-stroke: 0px;
}}
.hero-desc {{
    font-size: 0.88rem;
    color: var(--text2);
    max-width: 480px;
    line-height: 1.65;
    letter-spacing: 0.01em;
}}
.hero-grid-line {{
    position: absolute;
    right: 0; top: 0; bottom: 0;
    width: 42%;
    border-left: 1px solid var(--border);
    display: flex;
    align-items: center;
    justify-content: center;
}}
.hero-count-block {{
    text-align: center;
}}
.hero-count {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: 5rem;
    color: var(--text3);
    line-height: 1;
    letter-spacing: 0.05em;
}}
.hero-count-label {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.62rem;
    letter-spacing: 0.22em;
    text-transform: uppercase;
    color: var(--text3);
    margin-top: 0.3rem;
}}

/* ═══ SEARCH SECTION ═══ */
.search-section {{
    padding: 2rem 2.5rem;
    border-bottom: 1px solid var(--border);
    background: var(--bg2);
}}
.search-label {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.6rem;
    letter-spacing: 0.28em;
    text-transform: uppercase;
    color: var(--text3);
    margin-bottom: 0.6rem;
}}

/* ═══ APP GRID ═══ */
.apps-section {{
    padding: 2.5rem 2.5rem 1rem;
    background: var(--bg);
}}
.apps-label {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.6rem;
    letter-spacing: 0.28em;
    text-transform: uppercase;
    color: var(--text3);
    margin-bottom: 1.2rem;
    display: flex;
    align-items: center;
    gap: 0.8rem;
}}
.apps-label::after {{
    content: '';
    flex: 1;
    height: 1px;
    background: var(--border);
}}
.app-card {{
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 1.6rem 1.5rem 1.4rem;
    cursor: pointer;
    transition: border-color 0.2s, transform 0.2s, box-shadow 0.2s;
    position: relative;
    overflow: hidden;
}}
.app-card:hover {{
    border-color: var(--text);
    transform: translateY(-2px);
    box-shadow: 0 8px 32px rgba(0,0,0,0.12);
}}
.app-card-num {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: 2.8rem;
    color: var(--border2);
    line-height: 1;
    margin-bottom: 0.6rem;
}}
.app-card-title {{
    font-size: 0.95rem;
    font-weight: 600;
    color: var(--text);
    margin-bottom: 0.35rem;
    letter-spacing: -0.01em;
}}
.app-card-desc {{
    font-size: 0.75rem;
    color: var(--text2);
    line-height: 1.5;
    margin-bottom: 1.1rem;
}}
.app-card-tag {{
    display: inline-block;
    background: var(--tag-bg);
    color: var(--tag-text);
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.6rem;
    letter-spacing: 0.1em;
    padding: 0.22rem 0.6rem;
    border-radius: 3px;
    margin-right: 0.3rem;
    text-transform: uppercase;
}}
.app-card-arrow {{
    position: absolute;
    top: 1.4rem; right: 1.4rem;
    font-size: 1rem;
    color: var(--text3);
    transition: color 0.2s, transform 0.2s;
}}
.app-card:hover .app-card-arrow {{
    color: var(--text);
    transform: translate(2px, -2px);
}}

/* ═══ APP PANEL ═══ */
.app-panel {{
    background: var(--bg);
    padding: 2.5rem 2.5rem;
    border-top: 1px solid var(--border);
    animation: fadeUp 0.3s ease;
}}
@keyframes fadeUp {{
    from {{ opacity:0; transform: translateY(12px); }}
    to   {{ opacity:1; transform: translateY(0); }}
}}
.panel-header {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 1.8rem;
    padding-bottom: 1rem;
    border-bottom: 1px solid var(--border);
}}
.panel-title {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: 2rem;
    color: var(--text);
    letter-spacing: 0.04em;
}}
.panel-close {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.68rem;
    color: var(--text2);
    cursor: pointer;
    letter-spacing: 0.1em;
    border: 1px solid var(--border);
    padding: 0.3rem 0.8rem;
    border-radius: 3px;
    transition: all 0.15s;
}}
.panel-close:hover {{ border-color: var(--text); color: var(--text); }}

/* ═══ APP INNER ELEMENTS ═══ */
.app-h {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: 1.5rem;
    color: var(--text) !important;
    letter-spacing: 0.05em;
    margin-bottom: 0.2rem;
}}
.app-desc-sub {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.7rem;
    color: var(--text2);
    margin-bottom: 1.4rem;
    letter-spacing: 0.06em;
}}

/* ═══ STREAMLIT WIDGETS ═══ */
div[data-testid="stFileUploader"] {{
    background: var(--surface) !important;
    border: 1px dashed var(--border2) !important;
    border-radius: 6px !important;
    padding: 0.5rem !important;
    transition: border-color 0.2s !important;
}}
div[data-testid="stFileUploader"]:hover {{ border-color: var(--text) !important; }}
div[data-testid="stFileUploader"] label,
div[data-testid="stFileUploader"] p,
div[data-testid="stFileUploader"] span {{ color: var(--text) !important; }}

.stButton > button {{
    background: var(--btn-bg) !important;
    color: var(--btn-text) !important;
    border: none !important;
    border-radius: 4px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.72rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.08em !important;
    padding: 0.55rem 1.4rem !important;
    width: 100% !important;
    transition: opacity 0.15s !important;
    text-transform: uppercase !important;
}}
.stButton > button:hover {{ opacity: 0.85 !important; }}

.stDownloadButton > button {{
    background: var(--surface2) !important;
    color: var(--text) !important;
    border: 1px solid var(--border2) !important;
    border-radius: 4px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.72rem !important;
    letter-spacing: 0.08em !important;
    text-transform: uppercase !important;
    transition: border-color 0.15s !important;
}}
.stDownloadButton > button:hover {{
    border-color: var(--text) !important;
}}

/* ═══ METRIC CARD ═══ */
.metric-card {{
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 1.2rem 1rem;
    text-align: center;
}}
.metric-value {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: 2.2rem;
    color: var(--text) !important;
    letter-spacing: 0.04em;
    line-height: 1;
}}
.metric-label {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.58rem;
    color: var(--text2);
    letter-spacing: 0.18em;
    text-transform: uppercase;
    margin-top: 0.3rem;
}}

/* ═══ DATAFRAME ═══ */
.stDataFrame {{ border: 1px solid var(--border) !important; border-radius: 6px !important; }}

/* ═══ ALERTS ═══ */
.stAlert {{
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
}}
.stAlert p, .stAlert div {{ color: var(--text) !important; }}

/* ═══ TEXT ═══ */
p, li, span, label {{ color: var(--text); }}
h1, h2, h3, h4 {{ color: var(--text) !important; }}
hr {{ border-color: var(--border); }}
strong {{ color: var(--text); }}
.stSpinner > div {{ border-top-color: var(--text) !important; }}
.stMarkdown p {{ color: var(--text2); }}

/* ═══ SEARCH INPUT ═══ */
input[type="text"], .stTextInput input {{
    background: var(--surface) !important;
    color: var(--text) !important;
    border: 1px solid var(--border2) !important;
    border-radius: 4px !important;
    font-family: 'Instrument Sans', sans-serif !important;
    font-size: 0.85rem !important;
}}
input[type="text"]:focus, .stTextInput input:focus {{
    border-color: var(--text) !important;
    box-shadow: none !important;
    outline: none !important;
}}
.stTextInput label {{ color: var(--text2) !important; font-size: 0.72rem !important; }}

/* ═══ SELECTBOX ═══ */
.stSelectbox div[data-baseweb="select"] > div {{
    background: var(--surface) !important;
    border-color: var(--border2) !important;
    color: var(--text) !important;
    border-radius: 4px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.75rem !important;
}}

/* ═══ DIVIDER ═══ */
.section-divider {{
    height: 1px;
    background: var(--border);
    margin: 0 2.5rem;
}}

/* ═══ FOOTER ═══ */
.footer {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1.8rem 2.5rem;
    border-top: 1px solid var(--border);
    margin-top: 4rem;
    background: var(--bg2);
}}
.footer-logo {{
    font-family: 'Bebas Neue', sans-serif;
    font-size: 1.2rem;
    color: var(--text);
    letter-spacing: 0.1em;
}}
.footer-credit {{
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.65rem;
    color: var(--text2);
    letter-spacing: 0.18em;
    text-transform: uppercase;
}}
.footer-credit span {{
    color: var(--text);
    font-weight: 500;
}}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════
#  TICKER
# ════════════════════════════════════════
st.markdown("""
<div class="ticker-wrap">
  <span class="ticker-inner">
    &nbsp;&nbsp;&nbsp;&nbsp;UK PDF EXTRACTION SUITE
    &nbsp;&nbsp;·&nbsp;&nbsp; ESIC CHALLAN EXTRACTOR
    &nbsp;&nbsp;·&nbsp;&nbsp; TDS CHALLAN EXTRACTOR
    &nbsp;&nbsp;·&nbsp;&nbsp; EXPORT TO EXCEL
    &nbsp;&nbsp;·&nbsp;&nbsp; CREATED BY UDAY KIRAN
    &nbsp;&nbsp;·&nbsp;&nbsp; UK PDF EXTRACTION SUITE
    &nbsp;&nbsp;·&nbsp;&nbsp; ESIC CHALLAN EXTRACTOR
    &nbsp;&nbsp;·&nbsp;&nbsp; TDS CHALLAN EXTRACTOR
    &nbsp;&nbsp;·&nbsp;&nbsp; EXPORT TO EXCEL
    &nbsp;&nbsp;·&nbsp;&nbsp; CREATED BY UDAY KIRAN
  </span>
</div>
""", unsafe_allow_html=True)

# ════════════════════════════════════════
#  NAVBAR
# ════════════════════════════════════════
st.markdown("""
<div class="navbar">
    <div class="nav-logo">
        <div class="nav-logo-badge">U</div>
        UK
    </div>
    <div class="nav-links">
        <div class="nav-pill active">Explore ▾</div>
        <div class="nav-pill">Apps</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ════════════════════════════════════════
#  NAVBAR CONTROLS (mode + explore) via Streamlit
# ════════════════════════════════════════
with st.container():
    nc1, nc2, nc3, nc4, nc5 = st.columns([1, 1, 1, 3, 1])
    with nc1:
        if st.button("⬛ ESIC App", key="nav_esic"):
            st.session_state.active_app = "ESIC Challan Extractor"
            st.rerun()
    with nc2:
        if st.button("⬛ TDS App", key="nav_tds"):
            st.session_state.active_app = "TDS Challan Extractor"
            st.rerun()
    with nc5:
        mode_choice = st.selectbox(
            "", ["System", "Dark", "Light"],
            index=["System", "Dark", "Light"].index(st.session_state.color_mode),
            key="mode_sel",
            label_visibility="collapsed"
        )
        if mode_choice != st.session_state.color_mode:
            st.session_state.color_mode = mode_choice
            st.rerun()

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ════════════════════════════════════════
#  HERO
# ════════════════════════════════════════
st.markdown("""
<div class="hero-section">
    <div class="hero-eyebrow">PDF Extraction Suite &nbsp;/&nbsp; 2025</div>
    <div class="hero-title">UK</div>
    <div class="hero-desc">
        A minimal, powerful suite for extracting structured data from
        ESIC &amp; TDS challan PDFs — exported instantly to Excel.
    </div>
    <div class="hero-grid-line">
        <div class="hero-count-block">
            <div class="hero-count">02</div>
            <div class="hero-count-label">Apps Available</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ════════════════════════════════════════
#  SEARCH
# ════════════════════════════════════════
st.markdown('<div class="search-section"><div class="search-label">Search by app name</div></div>', unsafe_allow_html=True)

with st.container():
    sc1, sc2, sc3 = st.columns([1, 2, 1])
    with sc2:
        search = st.text_input(
            "", placeholder="🔍  Search apps…  e.g. ESIC, TDS",
            key="search_input",
            label_visibility="collapsed"
        )

# filter apps by search
filtered_apps = {k: v for k, v in APPS.items()
                 if not search or search.lower() in k.lower() or search.lower() in v.lower()}

# ════════════════════════════════════════
#  APP CARDS GRID
# ════════════════════════════════════════
st.markdown('<div class="apps-section"><div class="apps-label">My Apps</div></div>', unsafe_allow_html=True)

if not filtered_apps:
    st.markdown('<div style="padding:1rem 2.5rem; color:var(--text2); font-family:JetBrains Mono,monospace; font-size:0.78rem;">No apps match your search.</div>', unsafe_allow_html=True)
else:
    card_cols = st.columns(len(filtered_apps), gap="medium")
    app_names = list(filtered_apps.keys())
    for idx, col in enumerate(card_cols):
        if idx < len(app_names):
            aname = app_names[idx]
            adesc = filtered_apps[aname]
            tag = "ESIC" if "ESIC" in aname else "TDS"
            with col:
                st.markdown(f"""
                <div class="app-card">
                    <div class="app-card-num">0{idx+1}</div>
                    <div class="app-card-title">{aname}</div>
                    <div class="app-card-desc">{adesc}</div>
                    <span class="app-card-tag">{tag}</span>
                    <span class="app-card-tag">PDF → Excel</span>
                    <div class="app-card-arrow">↗</div>
                </div>
                """, unsafe_allow_html=True)
                if st.button(f"Open {aname}", key=f"open_{idx}"):
                    if st.session_state.active_app == aname:
                        st.session_state.active_app = None
                    else:
                        st.session_state.active_app = aname
                    st.rerun()

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ════════════════════════════════════════
#  APP PANELS (shown inline below cards)
# ════════════════════════════════════════

# ── ESIC LOGIC (unchanged) ──
ESIC_FIELDS = {
    "Employer's Code No": "Employer Code No",
    "Employer's Name": "Employer Name",
    "Challan Period": "Challan Period",
    "Challan Number": "Challan Number",
    "Challan Created Date": "Challan Created Date",
    "Challan Submitted Date": "Challan Submitted Date",
    "Amount Paid": "Amount Paid",
    "Transaction Number": "Transaction Number",
    "Transaction status": "Transaction Status",
}

def esic_extract_from_pdf(file_bytes):
    data = {}
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    for field, col in ESIC_FIELDS.items():
        pattern = re.escape(field) + r"[\s:]*([^\n]+)"
        match = re.search(pattern, text, re.IGNORECASE)
        data[col] = match.group(1).strip().rstrip("*").strip() if match else ""
    return data

def esic_create_excel(records):
    wb = Workbook(); ws = wb.active; ws.title = "ESIC Challans"
    headers = ["Source File"] + list(ESIC_FIELDS.values())
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", start_color="1A1A2E")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df2 = Font(name="Arial", size=10)
    af = PatternFill("solid", start_color="F2F2EF")
    ca = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = bdr
    ws.row_dimensions[1].height = 30
    for ri, record in enumerate(records, 2):
        fill = PatternFill("solid", start_color="FFFFFF") if ri % 2 == 0 else af
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=ri, column=ci, value=record.get(h, ""))
            cell.font = df2; cell.fill = fill; cell.alignment = ca; cell.border = bdr
        ws.row_dimensions[ri].height = 20
    col_widths = [30, 22, 28, 14, 22, 22, 22, 14, 22, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    tr = len(records) + 2
    ws.cell(row=tr, column=1, value="TOTAL").fill = PatternFill("solid", start_color="C0392B")
    ws.cell(row=tr, column=1).font = Font(name="Arial", bold=True, color="FFFFFF")
    ws.cell(row=tr, column=1).alignment = ca
    ac = headers.index("Amount Paid") + 1
    tc = ws.cell(row=tr, column=ac, value=f"=SUM({get_column_letter(ac)}2:{get_column_letter(ac)}{tr-1})")
    tc.font = Font(name="Arial", bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color="C0392B"); tc.alignment = ca; tc.border = bdr
    ws.freeze_panes = "A2"
    out = BytesIO(); wb.save(out); out.seek(0); return out

# ── TDS LOGIC (unchanged) ──
def tds_extract_value(text, label):
    for pattern in [rf"{re.escape(label)}\s*[:\-]\s*(.+)", rf"{re.escape(label)}\s+(.+)"]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m: return m.group(1).strip()
    return ""

def tds_clean_amount(val):
    val = val.replace("₹", "").replace(",", "").strip()
    m = re.search(r"[\d]+(?:\.\d+)?", val)
    return float(m.group()) if m else 0.0

def tds_extract_challan_data(pdf_file):
    with pdfplumber.open(pdf_file) as pdf:
        full_text = "".join((p.extract_text() or "") + "\n" for p in pdf.pages)
    data = {}
    for label, key in {
        "TAN":"TAN","Name":"Name","Assessment Year":"Assessment Year",
        "Financial Year":"Financial Year","Major Head":"Major Head",
        "Minor Head":"Minor Head","Nature of Payment":"Nature of Payment",
        "CIN":"CIN","Mode of Payment":"Mode of Payment","Bank Name":"Bank Name",
        "Bank Reference Number":"Bank Reference Number","Date of Deposit":"Date of Deposit",
        "BSR code":"BSR Code","Challan No":"Challan No","Tender Date":"Tender Date",
    }.items():
        data[key] = tds_extract_value(full_text, label)
    am = re.search(r"Amount \(in Rs\.\)\s*[:\-]?\s*₹?\s*([\d,]+)", full_text)
    data["Amount (Rs.)"] = tds_clean_amount(am.group(1)) if am else 0.0
    awm = re.search(r"Amount \(in words\)\s*[:\-]?\s*(.+)", full_text)
    data["Amount (in words)"] = awm.group(1).strip() if awm else ""
    for key, pat in {
        "Tax":r"A\s+Tax\s+₹?\s*([\d,]+)","Surcharge":r"B\s+Surcharge\s+₹?\s*([\d,]+)",
        "Cess":r"C\s+Cess\s+₹?\s*([\d,]+)","Interest":r"D\s+Interest\s+₹?\s*([\d,]+)",
        "Penalty":r"E\s+Penalty\s+₹?\s*([\d,]+)",
        "Fee u/s 234E":r"F\s+Fee under section 234E\s+₹?\s*([\d,]+)",
        "Total":r"Total \(A\+B\+C\+D\+E\+F\)\s+₹?\s*([\d,]+)",
    }.items():
        m = re.search(pat, full_text)
        data[key] = tds_clean_amount(m.group(1)) if m else 0.0
    im = re.search(r"ITNS No\.\s*[:\-]?\s*(\d+)", full_text)
    data["ITNS No."] = im.group(1).strip() if im else ""
    return data

def tds_create_excel(records):
    wb = Workbook(); ws = wb.active; ws.title = "TDS Challans"
    hfill = PatternFill("solid", start_color="1a1a2e", end_color="1a1a2e")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    sfill = PatternFill("solid", start_color="E8F4FD", end_color="E8F4FD")
    sfont = Font(bold=True, name="Arial", size=9, color="1a1a2e")
    dfont = Font(name="Arial", size=9)
    afill = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
    c = Alignment(horizontal="center", vertical="center")
    l = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="DEE2E6")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:T1")
    ws["A1"] = "TDS CHALLAN DETAILS — KAPSTON SERVICES LIMITED"
    ws["A1"].font = Font(bold=True, name="Arial", size=12, color="FFFFFF")
    ws["A1"].fill = hfill; ws["A1"].alignment = c; ws.row_dimensions[1].height = 28
    mh = ["S.No","ITNS No.","TAN","Name","Assessment Year","Financial Year",
          "Nature of Payment","CIN","Mode of Payment","Bank Name","Bank Ref. No.",
          "Date of Deposit","BSR Code","Challan No","Tender Date","Tax (Rs.)",
          "Surcharge (Rs.)","Cess (Rs.)","Interest (Rs.)","Penalty (Rs.)",
          "Fee u/s 234E (Rs.)","Total Amount (Rs.)"]
    ws.merge_cells("A2:A3"); ws.merge_cells("B2:B3")
    for ci, h in enumerate(mh, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = c; cell.border = bdr
        ws.cell(row=3, column=ci).font = sfont
        ws.cell(row=3, column=ci).fill = sfill
        ws.cell(row=3, column=ci).alignment = c
        ws.cell(row=3, column=ci).border = bdr
    ws.row_dimensions[2].height = 20; ws.row_dimensions[3].height = 16
    for i, rec in enumerate(records):
        row = i + 4
        fill = afill if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        vals = [i+1,rec.get("ITNS No.",""),rec.get("TAN",""),rec.get("Name",""),
                rec.get("Assessment Year",""),rec.get("Financial Year",""),
                rec.get("Nature of Payment",""),rec.get("CIN",""),
                rec.get("Mode of Payment",""),rec.get("Bank Name",""),
                rec.get("Bank Reference Number",""),rec.get("Date of Deposit",""),
                rec.get("BSR Code",""),rec.get("Challan No",""),rec.get("Tender Date",""),
                rec.get("Tax",0),rec.get("Surcharge",0),rec.get("Cess",0),
                rec.get("Interest",0),rec.get("Penalty",0),rec.get("Fee u/s 234E",0),rec.get("Total",0)]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = dfont; cell.fill = fill; cell.border = bdr
            cell.alignment = c if ci == 1 else l
            if ci >= 16: cell.number_format = '₹#,##0.00'
        ws.row_dimensions[row].height = 18
    tr = len(records) + 4
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=1).fill = hfill
    ws.cell(row=tr, column=1).font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    ws.cell(row=tr, column=1).alignment = c
    ws.merge_cells(f"A{tr}:O{tr}")
    for ci in range(16, 23):
        cl = get_column_letter(ci)
        cell = ws.cell(row=tr, column=ci, value=f"=SUM({cl}4:{cl}{tr-1})")
        cell.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
        cell.fill = hfill; cell.number_format = '₹#,##0.00'; cell.alignment = c; cell.border = bdr
    ws.row_dimensions[tr].height = 20
    for i, w in enumerate([5,8,14,28,14,12,18,26,14,14,18,14,10,10,12,14,14,10,10,10,14,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    out = BytesIO(); wb.save(out); out.seek(0); return out


# ── Render active app panel ──
if st.session_state.active_app == "ESIC Challan Extractor":
    st.markdown('<div class="app-panel">', unsafe_allow_html=True)
    ph1, ph2 = st.columns([8, 1])
    with ph1:
        st.markdown('<div class="panel-title">ESIC CHALLAN EXTRACTOR</div>', unsafe_allow_html=True)
    with ph2:
        if st.button("✕ Close", key="close_esic"):
            st.session_state.active_app = None
            st.rerun()

    st.markdown('<div class="app-desc-sub">Upload one or more ESIC Challan PDFs → extract fields → export to Excel</div>', unsafe_allow_html=True)

    esic_files = st.file_uploader(
        "Upload ESIC Challan PDFs",
        type=["pdf"], accept_multiple_files=True,
        help="You can upload multiple PDFs at once", key="esic_uploader"
    )
    if esic_files:
        st.markdown(f"**{len(esic_files)} file(s) uploaded**")
        records, errors = [], []
        for f in esic_files:
            try:
                r = esic_extract_from_pdf(f.read()); r["_filename"] = f.name; records.append(r)
            except Exception as e:
                errors.append(f"{f.name}: {e}")
        for err in errors: st.error(f"⚠️ {err}")
        if records:
            display_records = [{"Source File": r["_filename"], **{k: v for k, v in r.items() if k != "_filename"}} for r in records]
            st.markdown("### Preview")
            st.dataframe(pd.DataFrame(display_records), use_container_width=True)
            try:
                total_amt = sum(float(r.get("Amount Paid", 0) or 0) for r in display_records)
            except Exception:
                total_amt = 0.0
            st.markdown(f"**{len(records)} record(s) extracted** | Total Amount: ₹{total_amt:,.2f}")
            if st.button("⬇ Download Excel", key="esic_btn"):
                st.download_button(
                    label="📥 Save Excel File", data=esic_create_excel(display_records),
                    file_name="ESIC_Challans.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="esic_download"
                )
    else:
        st.info("Upload PDFs above to get started.")
    st.markdown('</div>', unsafe_allow_html=True)

elif st.session_state.active_app == "TDS Challan Extractor":
    st.markdown('<div class="app-panel">', unsafe_allow_html=True)
    ph1, ph2 = st.columns([8, 1])
    with ph1:
        st.markdown('<div class="panel-title">TDS CHALLAN EXTRACTOR</div>', unsafe_allow_html=True)
    with ph2:
        if st.button("✕ Close", key="close_tds"):
            st.session_state.active_app = None
            st.rerun()

    st.markdown('<div class="app-desc-sub">Upload ITNS 281 challan receipts → extract all data → export to Excel</div>', unsafe_allow_html=True)

    tds_files = st.file_uploader(
        "Upload challan PDF files", type=["pdf"], accept_multiple_files=True,
        help="Upload one or more ITNS 281 TDS challan PDF files", key="tds_uploader"
    )
    if tds_files:
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f'<div class="metric-card"><div class="metric-value">{len(tds_files)}</div><div class="metric-label">Files Uploaded</div></div>', unsafe_allow_html=True)
        records, errors = [], []
        with st.spinner("Extracting data from PDFs..."):
            for f in tds_files:
                try:
                    d = tds_extract_challan_data(f); d["_filename"] = f.name; records.append(d)
                except Exception as e:
                    errors.append((f.name, str(e)))
        for fname, err in errors: st.error(f"❌ {fname}: {err}")
        if records:
            total_amount = sum(r.get("Total", 0) for r in records)
            with col2:
                st.markdown(f'<div class="metric-card"><div class="metric-value">{len(records)}</div><div class="metric-label">Extracted</div></div>', unsafe_allow_html=True)
            with col3:
                st.markdown(f'<div class="metric-card"><div class="metric-value">₹{total_amount:,.0f}</div><div class="metric-label">Total TDS</div></div>', unsafe_allow_html=True)
            st.markdown("### 📊 Extracted Data Preview")
            preview_cols = ["Nature of Payment","CIN","Challan No","Date of Deposit","BSR Code","Tax","Surcharge","Cess","Interest","Penalty","Fee u/s 234E","Total"]
            df = pd.DataFrame(records)
            df.insert(0, "S.No", range(1, len(df)+1))
            df["File"] = df["_filename"]
            display_cols = ["S.No", "File"] + [c for c in preview_cols if c in df.columns]
            st.dataframe(df[display_cols], use_container_width=True, hide_index=True)
            st.markdown("### 💾 Export to Excel")
            st.download_button(
                label="⬇️ Download Excel File", data=tds_create_excel(records),
                file_name="TDS_Challans.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="tds_download"
            )
    else:
        st.info("👆 Upload one or more TDS challan PDF files to get started.")
        st.markdown("**Supported format:** ITNS 281 Challan Receipts from Income Tax Department")
    st.markdown('</div>', unsafe_allow_html=True)

# ════════════════════════════════════════
#  FOOTER
# ════════════════════════════════════════
st.markdown("""
<div class="footer">
    <div class="footer-logo">UK</div>
    <div class="footer-credit">Created by <span>Uday Kiran</span></div>
    <div class="footer-credit">PDF Extraction Suite</div>
</div>
""", unsafe_allow_html=True)
